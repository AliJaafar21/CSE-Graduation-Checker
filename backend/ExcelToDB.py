from openpyxl import load_workbook
from src.database import db, Course


def populateTheDatabase():
	print("Populating The Database")

	workbook = load_workbook("COURSES.xlsx")
	for sheetName in workbook.sheetnames:
		print(f" Processing {sheetName} Sheet")
		sheet = workbook[sheetName]			
		for rowIndex in range(1, sheet.max_row + 1):	
			name = sheet.cell(row=rowIndex, column=1).value
			title = sheet.cell(row=rowIndex, column=2).value
			tag1 = sheet.cell(row=rowIndex, column=3).value
			tag2 = sheet.cell(row=rowIndex, column=4).value
			tag3 = sheet.cell(row=rowIndex, column=5).value
				
			if(Course.query.filter_by(name=name)).first():
				print(f"  {name} already exits")
				continue

			course = Course(name=name, title=title, tag1=tag1, tag2=tag2, tag3=tag3, equivalent_to=None)
			db.session.add(course)
	db.session.commit()
	workbook.close()


def setEquivalencies():
	print("Setting Equivalencies")
	workbook = load_workbook("EQUIVALENCIES.xlsx")
	sheet = workbook["Equivalencies"]
	for rowIndex in range(1, sheet.max_row + 1):
		course1Name = sheet.cell(row=rowIndex, column=1).value
		course2Name = sheet.cell(row=rowIndex, column=2).value
		
		course1 = Course.query.filter_by(name=course1Name).first()
		if(course1 is None):
			print(f" {course1Name} doesn't exist")
			continue
		
		course2 = Course.query.filter_by(name=course2Name).first()
		if(course2 is None):
			print(f" {course2Name} doesn't exist")
			continue

		course1.equivalent_to = course2Name
		course2.equivalent_to = course1Name
	db.session.commit()
	workbook.close()


db.create_all()
populateTheDatabase()
setEquivalencies()